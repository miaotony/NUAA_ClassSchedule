#!/usr/bin/python3
# -*- coding:utf-8 -*-
"""
generateXLSX  生成及导出.xlsx表格文件

@Author: ZegWe
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
import os


def create_xls(lessons, semester_year, semester, stuID):
    """
    生成`.xlsx`表格文件
    :param lessons: {list} 课程列表
    :param semester_year: {str} 学年
    :param semester: {str} 学期
    :param stuID: {str} 学号
    :return: book {class Workbook} 表格
    """
    book = Workbook()
    sheet = book.create_sheet('classTable', 0)
    alignment = Alignment(
        horizontal='center',
        vertical='center',
        shrink_to_fit=True,
        wrap_text=True,
    )
    sheet.merge_cells('A1:H1')
    # print('1\n')
    sheet.cell(1, 1, 'Semester year: ' + semester_year +
               '  Semester: ' + semester + '  ID: ' + stuID).alignment = alignment
    sheet.column_dimensions['A'].width = 3
    sheet.row_dimensions[1].height = 20
    sheet.row_dimensions[2].height = 20
    # print('1\n')
    for i in range(1, 12):
        sheet.cell(i + 2, 1, i).alignment = alignment
        sheet.row_dimensions[i + 2].height = 60
    weekday = ['一', '二', '三', '四', '五', '六', '天']
    ab = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    # print('2\n')
    for i in range(0, 7):
        sheet.cell(2, i + 2, '星期' + weekday[i]).alignment = alignment
        sheet.column_dimensions[ab[i + 1]].width = 20
    # print('xls created\n')
    for lesson in lessons:
        unit = int(lesson.course_unit[0])
        day = int(lesson.day_of_week)
        rag = ab[day] + str(unit + 2) + ':' + ab[day] + \
              str(unit + 1 + len(lesson.course_unit))
        # print(rag)
        sheet.merge_cells(rag)
        if sheet.cell(unit + 2, day + 1).value:  # Exist other lesson(s)
            sheet.cell(unit + 2, day + 1).value = str(sheet.cell(unit + 2,
                                                                 day + 1).value) + '\n\n' + lesson.str_for_print
        else:
            sheet.cell(unit + 2, day + 1).value = lesson.str_for_print
            sheet.cell(unit + 2, day + 1).alignment = alignment
            sheet.cell(unit + 2, day + 1).fill = PatternFill('solid', 'FFFF00')
            # stl = Side('double')
            # sheet.cell(unit+2, day+1).border = Border(stl,stl,stl,stl)
    return book


def export_xls(xls, semester_year, semester, stuID):
    """
    导出表格文件
    :param xls: {class Workbook} 表格
    :param semester_year: {str} 学年
    :param semester: {str} 学期
    :param stuID: {str} 学号
    :return: None
    """
    try:
        filename = 'NUAAiCal-Data/NUAA-curriculum-' + \
                   semester_year + '-' + semester + '-' + stuID + '.xlsx'
        xls.save(os.path.abspath(filename))
        print('表格文件已导出到\"' + os.path.abspath(filename) + '\"。')
    except Exception as e:
        print('ERROR! 导出到xlsx文件错误')
        print(e)
